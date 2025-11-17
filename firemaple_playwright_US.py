# -*- coding: utf-8 -*-
"""
firemaple_playwright_us.py
é€šè¿‡é“¾æ¥æŠ“å– Amazon US å•†å“ä¿¡æ¯ï¼ˆæ‰‹åŠ¨ä¿®æ”¹åœ°å€ç‰ˆï¼‰
æ–°å¢ï¼š
  1) â€œäº§å“å›¾ç‰‡â€ åˆ—ï¼šä¸»å›¾ URLï¼ˆCSV ç¬¬ä¸€åˆ—ï¼‰
  2) ç”Ÿæˆ .xlsxï¼Œé¦–åˆ—åµŒå…¥ä¸»å›¾ç¼©ç•¥å›¾
  3) åº—é“ºåç§°è‡ªåŠ¨æ¸…æ´—ï¼ˆå»æ‰ Sold by/Ships from å‰ç¼€ã€å»é‡ç­‰ï¼‰

è¾“å‡ºå­—æ®µï¼š
äº§å“å›¾ç‰‡ / é“¾æ¥ / äºšé©¬é€ŠASIN / ä»·æ ¼ / ç±»ç›®&æ’å / è¯„åˆ† / åº—é“ºåç§° / æ˜¯å¦FBA / reviewæ•°é‡ / reviewæƒ…å†µ
"""

import asyncio
import re
import os
import io
import random
import requests
import pandas as pd
from bs4 import BeautifulSoup
from tqdm import tqdm
from PIL import Image as PILImage
from playwright.async_api import async_playwright

# ============ é€šç”¨å·¥å…· ============
def clean_text(txt):
    if not txt:
        return "â€”"
    return re.sub(r"\s+", " ", txt).strip()


# ============ ASIN & FBA è¾…åŠ©å‡½æ•°ï¼ˆé€šç”¨ç‰ˆï¼Œé€‚ç”¨ US/UK ç­‰ï¼‰ ============
def get_asin_from_url(url):
    """
    ä» URL ä¸­æå– ASINï¼Œå…¼å®¹ /dp/ã€/gp/product/ã€/product/ ç­‰å¤šç§å½¢å¼
    """
    patterns = [
        r"/dp/([A-Z0-9]{10})",
        r"/gp/product/([A-Z0-9]{10})",
        r"/product/([A-Z0-9]{10})",
    ]
    for pat in patterns:
        m = re.search(pat, url)
        if m:
            return m.group(1)
    return None


def get_asin_from_page(soup):
    """
    ä»é¡µé¢è¯¦æƒ…è¡¨æ ¼ / detail bullets ä¸­æå– ASIN
    """
    # äº§å“è¯¦æƒ…è¡¨æ ¼é‡Œæ‰¾ ASIN
    tables = soup.select(
        "table#productDetails_detailBullets_sections1, "
        "table#productDetails_techSpec_section_1, "
        "table.prodDetTable"
    )
    for table in tables:
        for row in table.select("tr"):
            header = row.select_one("th")
            if not header:
                continue
            label = header.get_text(strip=True)
            if label == "ASIN":
                val = row.select_one("td")
                if val:
                    asin = val.get_text(strip=True)
                    if re.fullmatch(r"[A-Z0-9]{10}", asin):
                        return asin

    # detailBullets_feature_div åŒºåŸŸå…œåº•
    bullets = soup.select("#detailBullets_feature_div li")
    for li in bullets:
        label = li.select_one("span.a-text-bold")
        if label and "ASIN" in label.get_text():
            text = li.get_text(" ", strip=True)
            m = re.search(r"([A-Z0-9]{10})", text)
            if m:
                return m.group(1)

    return None


def get_asin(url, soup):
    """
    ç»¼åˆ URL + é¡µé¢ä¸¤ç§æ–¹å¼è·å– ASIN
    """
    asin = get_asin_from_url(url)
    if asin:
        return asin
    asin = get_asin_from_page(soup)
    return asin if asin else "â€”"


def detect_fba(soup, ships_from_text, seller_text):
    """
    åˆ¤æ–­æ˜¯å¦ FBAï¼š
    - ç»¼åˆ Ships from / Sold by / merchant-info / buybox ç­‰åŒºåŸŸ
    - è¯†åˆ«ï¼šfulfilled/dispatches/dispatched/ships/delivered/sold ... by/from Amazon
    è¿”å› "æ˜¯" æˆ– "å¦"
    """
    text_blocks = []

    if ships_from_text and ships_from_text != "â€”":
        text_blocks.append(ships_from_text)

    if seller_text and seller_text != "â€”":
        text_blocks.append(seller_text)

    el = soup.select_one("#merchant-info")
    if el:
        text_blocks.append(el.get_text(" ", strip=True))

    el = soup.select_one("#tabular-buybox")
    if el:
        text_blocks.append(el.get_text(" ", strip=True))

    el = soup.select_one("#shipsFromSoldBy_feature_div")
    if el:
        text_blocks.append(el.get_text(" ", strip=True))

    el = soup.select_one("#desktop_buybox")
    if el:
        text_blocks.append(el.get_text(" ", strip=True))

    blob = " ".join(text_blocks).lower()

    # 1) ç²¾å‡†åŒ¹é…å„ç§å¸¸è§å†™æ³•
    patterns = [
        r"fulfilled\s+by\s+amazon",
        r"dispatch(?:es|ed)?\s+from\s+amazon",   # dispatches from / dispatched from Amazon
        r"ships?\s+from\s+amazon",
        r"delivered\s+by\s+amazon",
        r"sold\s+by\s+amazon",
    ]
    for pat in patterns:
        if re.search(pat, blob):
            return "æ˜¯"

    # 2) æ¨¡ç³Šï¼šåŒ…å« amazon ä¸”é™„è¿‘æœ‰ dispatch/ship/fulfil/prime ç­‰å­—æ ·
    if "amazon" in blob and any(
        kw in blob for kw in ["dispatch", "ship", "fulfil", "fulfill", "prime", "delivery"]
    ):
        return "æ˜¯"

    # 3) å…œåº•ï¼šships_from å­—æ®µé‡Œæœ¬èº«å°±å†™äº† amazon ä¹Ÿè§†ä¸º FBA
    if ships_from_text and "amazon" in ships_from_text.lower():
        return "æ˜¯"

    return "å¦"


# ============ æ‰‹åŠ¨è®¾ç½®åœ°å€ï¼ˆç¾å›½ï¼‰ ============
async def set_us_delivery_address(page):
    print("ğŸ”¹ æ­£åœ¨æ‰“å¼€ Amazon US é¦–é¡µï¼Œè¯·æ‰‹åŠ¨å°†æ”¶è´§åœ°å€ä¿®æ”¹ä¸ºç¾å›½ï¼ˆå»ºè®®é‚®ç¼– 10001ï¼‰...")
    print("   ä¿®æ”¹å®Œæˆåè¿”å›ç»ˆç«¯æŒ‰ Enter ç»§ç»­ã€‚")
    await page.goto("https://www.amazon.com/", timeout=60000, wait_until="domcontentloaded")
    await page.wait_for_timeout(2000)
    input("ğŸ‘‰ è¯·æ‰‹åŠ¨ä¿®æ”¹åœ°å€å®ŒæˆåæŒ‰ Enter é”®ç»§ç»­æŠ“å–...")


# ============ æŠ“å–å•ä¸ªå•†å“ ============
async def fetch_product(page, url):
    """æ‰“å¼€å•†å“é¡µå¹¶è§£æå­—æ®µï¼ˆå«ä¸»å›¾ URLï¼›åº—å/FBAé€»è¾‘ï¼‰"""
    try:
        await page.goto(url, timeout=60000, wait_until="domcontentloaded")
        await page.wait_for_selector("#productTitle", timeout=30000)
        await page.evaluate("window.scrollBy(0, 400)")
        await page.wait_for_timeout(1000)
        html = await page.content()
        soup = BeautifulSoup(html, "lxml")

        data = {}

        # ---------- äº§å“ä¸»å›¾ ----------
        img_url = None
        img_el = soup.select_one("#landingImage") or soup.select_one("#imgTagWrapperId img")
        if img_el and img_el.get("src"):
            img_url = img_el.get("src")
        if not img_url:
            thumb = soup.select_one("#altImages img, .imageThumbnail img")
            if thumb and thumb.get("src"):
                img_url = thumb.get("src")
        data["äº§å“å›¾ç‰‡"] = img_url if img_url else "â€”"

        # ---------- å•†å“é“¾æ¥ ----------
        data["é“¾æ¥"] = url

        # ---------- ASIN ----------
        data["äºšé©¬é€ŠASIN"] = get_asin(url, soup)

        # ---------- ä»·æ ¼ï¼ˆç¾å…ƒ $ï¼‰ ----------
        price = None
        for sel in [
            "#corePrice_feature_div .a-price .a-offscreen",
            "#apex_desktop .a-price .a-offscreen",
            "#corePrice_desktop_feature_div .a-price .a-offscreen",
            "#price_inside_buybox",
            "span.a-price .a-offscreen",
        ]:
            el = soup.select_one(sel)
            if el and "$" in el.get_text():
                price = el.get_text(strip=True)
                break
        if not price:
            for el in soup.select("span.a-offscreen"):
                parent = el.find_parent()
                pid = parent.get("id") if parent else ""
                if pid and re.search(r"(installment|emi)", pid, re.I):
                    continue
                txt = el.get_text(strip=True)
                if "$" in txt and re.search(r"\$\s?\d", txt) and len(txt) < 24:
                    price = txt
                    break
        if not price:
            whole = soup.select_one("span.a-price-whole")
            frac = soup.select_one("span.a-price-fraction")
            sym = soup.select_one("span.a-price-symbol")
            if whole:
                price = (sym.get_text(strip=True) if sym else "$") + whole.get_text(strip=True)
                if frac:
                    price += "." + frac.get_text(strip=True)
        data["ä»·æ ¼"] = clean_text(price)

        # ---------- è¯„åˆ† ----------
        rating_el = (
            soup.select_one("span[data-hook='rating-out-of-text']")
            or soup.select_one("i[data-hook='average-star-rating'] span")
            or soup.select_one("span.a-icon-alt")
        )
        data["è¯„åˆ†"] = clean_text(rating_el.get_text(strip=True) if rating_el else None)

        # ---------- review æ•°é‡ ----------
        rc_el = (
            soup.select_one("#acrCustomerReviewText")
            or soup.select_one("span#acrCustomerReviewText")
            or soup.select_one("[data-hook='total-review-count']")
            or soup.select_one("#acrPopover .a-size-base")
        )
        data["ratingæ•°é‡"] = clean_text(rc_el.get_text(strip=True) if rc_el else None)

        # ---------- åº—é“ºåç§° + æ˜¯å¦FBA ----------
        seller = "â€”"
        ships_from = "â€”"

        # æ–°ç‰ˆ tabular buybox
        for block in soup.select("#tabular-buybox .tabular-buybox-container, #tabular-buybox .tabular-buybox-text-row"):
            label_el = block.select_one(".tabular-buybox-label")
            text_el  = block.select_one(".tabular-buybox-text")
            if not label_el or not text_el:
                continue
            label = label_el.get_text(strip=True).lower()
            value = clean_text(text_el.get_text(strip=True))
            if "sold" in label and seller == "â€”":
                seller = value
            elif "ships" in label and ships_from == "â€”":
                ships_from = value

        # æ—§å¼ä¸¤è¡Œæ–‡æœ¬
        if seller == "â€”" or ships_from == "â€”":
            for box_sel in ["#shipsFromSoldBy_feature_div", "#desktop_buybox", "#rightCol", "#buybox_feature_div"]:
                box = soup.select_one(box_sel)
                if not box:
                    continue
                # Ships from
                if ships_from == "â€”":
                    lab = box.find(string=re.compile(r'^\s*Ships\s*from\s*$', re.I))
                    if lab:
                        row = lab.find_parent() or box
                        cand = row.find_next(lambda tag: tag.name in ["a", "span", "div"] and clean_text(tag.get_text()))
                        if cand:
                            val = clean_text(cand.get_text())
                            if val.lower() != "ships from":
                                ships_from = val
                # Sold by
                if seller == "â€”":
                    lab = box.find(string=re.compile(r'^\s*Sold\s*by\s*$', re.I))
                    if lab:
                        row = lab.find_parent() or box
                        cand = row.find_next(lambda tag: tag.name in ["a", "span", "div"] and clean_text(tag.get_text()))
                        if cand:
                            val = clean_text(cand.get_text())
                            if val.lower() != "sold by":
                                seller = val

                # å—å†…å…œåº•
                if ships_from == "â€”":
                    m1 = re.search(r"Ships\s*from\s+([A-Za-z0-9 &\-]+)", box.get_text(" ", strip=True), re.I)
                    if m1:
                        ships_from = clean_text(m1.group(1))
                if seller == "â€”":
                    m2 = re.search(r"Sold\s*by\s+(.+?)(?:\s+and|\s+\.|$)", box.get_text(" ", strip=True), re.I)
                    if m2:
                        seller = clean_text(m2.group(1))

        # merchant-info å…œåº•
        if seller == "â€”":
            mi = soup.select_one("#merchant-info")
            if mi:
                m = re.search(r"Sold\s*by\s+(.+?)(?:\s+and|\s+\.|$)", mi.get_text(" ", strip=True), re.I)
                if m:
                    seller = clean_text(m.group(1))

        data["åº—é“ºåç§°"] = seller

        # æ˜¯å¦FBAï¼ˆUS ä¹Ÿé€šç”¨ï¼‰
        data["æ˜¯å¦FBA"] = detect_fba(soup, ships_from, seller)

        # ---------- ç±»ç›®&æ’å ----------
        bsr = "â€”"
        for sel in ["#detailBullets_feature_div", "#productDetails_detailBullets_sections1", "#prodDetails"]:
            node = soup.select_one(sel)
            if not node:
                continue
            text = node.get_text(" ", strip=True)
            mm = re.search(r"Best\s*Sellers?\s*Rank\s*:?\s*(.+?)(?:Date First Available|Customer Reviews|ASIN|$)", text, flags=re.I)
            if mm:
                bsr = clean_text(mm.group(1))
                break
        if bsr == "â€”":
            crumbs = [a.get_text(strip=True) for a in soup.select("#wayfinding-breadcrumbs_feature_div a")]
            if crumbs:
                bsr = " / ".join([c for c in crumbs if c])
        data["ç±»ç›®&æ’å"] = bsr

        # ---------- review æƒ…å†µ ----------
        rv = (
            soup.select_one("div[data-hook='review'] span[data-hook='review-title'] span")
            or soup.select_one("div[data-hook='review'] span[data-hook='review-body'] span")
        )
        if rv:
            txt = rv.get_text(strip=True)
            data["reviewæƒ…å†µ"] = clean_text(txt[:120] + ("..." if len(txt) > 120 else ""))
        else:
            data["reviewæƒ…å†µ"] = "â€”"

        return data

    except Exception as e:
        print(f"[ERROR] {url} æŠ“å–å¤±è´¥ï¼š{e}")
        return None


# ============ æœ€ç»ˆç®€åŒ–+å»é‡ç‰ˆåº—é“ºåç§°æ¸…æ´—æ¨¡å— ============
def normalize_seller_name(name: str) -> str:
    """
    åº—é“ºåç§°æ¸…æ´—é€»è¾‘ï¼š
    - å»é™¤å‰åç©ºæ ¼
    - å¦‚æœåŒ…å« "Sold by"ï¼ˆä¸åŒºåˆ†å¤§å°å†™ï¼‰ï¼Œæˆªæ–­ä¿ç•™å‰éƒ¨åˆ†
    - å»æ‰é‡å¤å­ä¸²ï¼ˆå¦‚ "Cool Hand Store Cool Hand Store" â†’ "Cool Hand Store"ï¼‰
    """
    if not name or name == "â€”":
        return "â€”"

    s = name.strip()
    # é‡åˆ° Sold by å°±æˆªæ–­
    m = re.search(r"(?i)\bSold\s*by\b", s)
    if m:
        s = s[:m.start()]

    # å»é™¤å¤šä½™ç©ºæ ¼å’Œæ ‡ç‚¹
    s = s.strip(" .-â€“")

    # åˆ¤æ–­é‡å¤ï¼ˆæ•´ä¸²é‡å¤ä¸¤éçš„æƒ…å†µï¼‰
    parts = s.split()
    half = len(parts) // 2
    if len(parts) % 2 == 0 and parts[:half] == parts[half:]:
        s = " ".join(parts[:half])

    return s if s else "â€”"


def apply_seller_cleanup(rows):
    """å°±åœ°æ¸…æ´— rows é‡Œçš„â€œåº—é“ºåç§°â€å­—æ®µ"""
    for r in rows:
        if "åº—é“ºåç§°" in r:
            r["åº—é“ºåç§°"] = normalize_seller_name(r.get("åº—é“ºåç§°", "â€”"))


# ============ ç”Ÿæˆå¸¦å›¾ç‰‡çš„ Excel ============
def save_xlsx_with_images(rows, xlsx_path="firemaple_playwright_us.xlsx"):
    """
    å°†æŠ“å–ç»“æœå†™å…¥ .xlsxï¼Œå¹¶æŠŠâ€œäº§å“å›¾ç‰‡â€åµŒå…¥é¦–åˆ—ç¼©ç•¥å›¾ã€‚
    ä¼šå°è¯•ä¸‹è½½å›¾ç‰‡ï¼Œå¤±è´¥åˆ™ç•™ç©ºã€‚
    """
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Fire-Maple US"

    headers = ["äº§å“å›¾ç‰‡","é“¾æ¥","äºšé©¬é€ŠASIN","ä»·æ ¼","ç±»ç›®&æ’å","è¯„åˆ†","åº—é“ºåç§°","æ˜¯å¦FBA","ratingæ•°é‡","reviewæƒ…å†µ"]
    ws.append(headers)

    # è®¾ç½®åˆ—å®½ï¼Œè¡Œé«˜ï¼ˆé¦–åˆ—æ”¾ç¼©ç•¥å›¾ï¼‰
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 42
    for col_idx in range(3, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20

    # ä¸‹è½½å›¾ç‰‡ç”¨çš„ç®€å• headers
    http_headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120 Safari/537.36"
    }

    row_idx = 2
    for row in rows:
        # å…ˆå†™æ–‡æœ¬æ•°æ®ï¼ˆå›¾ç‰‡åˆ—ç•™ç©ºï¼Œç¨åæ’å…¥ï¼‰
        ws.append([
            "",  # å›¾ç‰‡ç¨åæ’å…¥
            row.get("é“¾æ¥",""),
            row.get("äºšé©¬é€ŠASIN",""),
            row.get("ä»·æ ¼",""),
            row.get("ç±»ç›®&æ’å",""),
            row.get("è¯„åˆ†",""),
            row.get("åº—é“ºåç§°",""),
            row.get("æ˜¯å¦FBA",""),
            row.get("ratingæ•°é‡",""),
            row.get("reviewæƒ…å†µ",""),
        ])

        # æ‹‰å–å›¾ç‰‡å¹¶ç¼©ç•¥
        img_url = row.get("äº§å“å›¾ç‰‡")
        if img_url and img_url != "â€”":
            try:
                r = requests.get(img_url, headers=http_headers, timeout=10)
                r.raise_for_status()
                img_bytes = io.BytesIO(r.content)
                with PILImage.open(img_bytes) as im:
                    im = im.convert("RGB")
                    im.thumbnail((120, 120))  # æ§åˆ¶ç¼©ç•¥å›¾å¤§å°
                    buf = io.BytesIO()
                    im.save(buf, format="JPEG", quality=85)
                    buf.seek(0)
                xl_img = XLImage(buf)
                xl_img.width, xl_img.height = im.size
                anchor = f"A{row_idx}"
                ws.add_image(xl_img, anchor)
                ws.row_dimensions[row_idx].height = 95  # è¡Œé«˜ç¨å¾®å¤§ä¸€ç‚¹
            except Exception:
                # ä¸‹è½½å¤±è´¥å°±ç•™ç©º
                pass

        row_idx += 1

    from openpyxl.styles import Alignment
    for col in "BCDEFGHIJ":
        for r in range(1, row_idx):
            ws[f"{col}{r}"].alignment = Alignment(vertical="center", wrap_text=True)

    wb.save(xlsx_path)
    print(f"[DONE] å·²ç”Ÿæˆå¸¦å›¾ç‰‡çš„ Excelï¼š{xlsx_path}")


# ============ ä¸»æµç¨‹ ============
async def main():
    # è¯»å–é“¾æ¥ï¼ˆç¾å›½ç«™ amazon.com é“¾æ¥ï¼‰
    with open("urls.txt", "r", encoding="utf-8") as f:
        urls = [line.strip() for line in f if line.strip()]

    results = []
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        context = await browser.new_context(locale="en-US", viewport={"width": 1280, "height": 900})
        page = await context.new_page()

        # æ‰‹åŠ¨è®¾ç½®ç¾å›½æ”¶è´§åœ°å€
        await set_us_delivery_address(page)

        for url in tqdm(urls, desc="æŠ“å–è¿›åº¦", unit="item"):
            data = await fetch_product(page, url)
            if data:
                results.append(data)
            await asyncio.sleep(2 + (random.random() * 2))

        await browser.close()

    # åº—é“ºåç§°æ¸…æ´—
    apply_seller_cleanup(results)

    # è¾“å‡º CSV
    if results:
        df = pd.DataFrame(
            results,
            columns=[
                "äº§å“å›¾ç‰‡",
                "é“¾æ¥",
                "äºšé©¬é€ŠASIN",
                "ä»·æ ¼",
                "ç±»ç›®&æ’å",
                "è¯„åˆ†",
                "åº—é“ºåç§°",
                "æ˜¯å¦FBA",
                "ratingæ•°é‡",
                "reviewæƒ…å†µ",
            ],
        )
        csv_path = "firemaple_playwright_us.csv"
        xlsx_path = "firemaple_playwright_us.xlsx"
        df.to_csv(csv_path, index=False, encoding="utf-8-sig")
        print(f"[DONE] å…±ä¿å­˜ {len(df)} æ¡åˆ° CSVï¼š{csv_path}")

        # ç”Ÿæˆå¸¦å›¾ç‰‡çš„ Excel
        save_xlsx_with_images(results, xlsx_path=xlsx_path)
    else:
        print("[ERROR] æ²¡æœ‰æˆåŠŸæŠ“å–åˆ°ä»»ä½•å•†å“ä¿¡æ¯ã€‚")


if __name__ == "__main__":
    asyncio.run(main())
