# -*- coding: utf-8 -*-
"""
firemaple_playwright.py
通过链接抓取 Amazon AU 商品信息（手动修改地址版）
新增：
  1) “产品图片” 列：主图 URL（CSV 第一列）
  2) 生成 .xlsx，首列嵌入主图缩略图
  3) 店铺名称自动清洗（去掉 Sold by/Ships from 前缀、去重等）

输出字段：
产品图片 / 链接 / 亚马逊ASIN / 价格 / 类目&排名 / 评分 / 店铺名称 / 是否FBA / review数量 / review情况
"""

import asyncio
import re
import os
import io
import random
import requests
import pandas as pd
from bs4 import BeautifulSoup
from tqdm import tqdm
from PIL import Image as PILImage
from playwright.async_api import async_playwright

# ============ 通用工具 ============
def clean_text(txt):
    if not txt:
        return "—"
    return re.sub(r"\s+", " ", txt).strip()

# ============ 手动设置地址 ============
async def set_au_delivery_address(page):
    print("🔹 正在打开 Amazon AU 首页，请手动将收货地址修改为澳洲（建议邮编 2000）...")
    print("   修改完成后返回终端按 Enter 继续。")
    await page.goto("https://www.amazon.com.au/", timeout=60000, wait_until="domcontentloaded")
    await page.wait_for_timeout(2000)
    input("👉 请手动修改地址完成后按 Enter 键继续抓取...")

# ============ 抓取单个商品 ============
async def fetch_product(page, url):
    """打开商品页并解析字段（含主图 URL；店名/FBA沿用稳定逻辑）"""
    try:
        await page.goto(url, timeout=60000, wait_until="domcontentloaded")
        await page.wait_for_selector("#productTitle", timeout=30000)
        await page.evaluate("window.scrollBy(0, 400)")
        await page.wait_for_timeout(1000)
        html = await page.content()
        soup = BeautifulSoup(html, "lxml")

        data = {}

        # ---------- 产品主图 ----------
        img_url = None
        img_el = soup.select_one("#landingImage") or soup.select_one("#imgTagWrapperId img")
        if img_el and img_el.get("src"):
            img_url = img_el.get("src")
        if not img_url:
            thumb = soup.select_one("#altImages img, .imageThumbnail img")
            if thumb and thumb.get("src"):
                img_url = thumb.get("src")
        data["产品图片"] = img_url if img_url else "—"

        # ---------- 商品链接 ----------
        data["链接"] = url

        # ---------- ASIN ----------
        m = re.search(r"/dp/([A-Z0-9]{10})", url)
        data["亚马逊ASIN"] = m.group(1) if m else "—"

        # ---------- 价格 ----------
        price = None
        for sel in [
            "#corePrice_feature_div .a-price .a-offscreen",
            "#apex_desktop .a-price .a-offscreen",
            "#corePrice_desktop_feature_div .a-price .a-offscreen",
            "#price_inside_buybox",
            "span.a-price .a-offscreen",
        ]:
            el = soup.select_one(sel)
            if el and "$" in el.get_text():
                price = el.get_text(strip=True)
                break
        if not price:
            for el in soup.select("span.a-offscreen"):
                parent = el.find_parent()
                pid = parent.get("id") if parent else ""
                if pid and re.search(r"(installment|emi)", pid, re.I):
                    continue
                txt = el.get_text(strip=True)
                if "$" in txt and re.search(r"\$\s?\d", txt) and len(txt) < 24:
                    price = txt
                    break
        if not price:
            whole = soup.select_one("span.a-price-whole")
            frac = soup.select_one("span.a-price-fraction")
            sym = soup.select_one("span.a-price-symbol")
            if whole:
                price = (sym.get_text(strip=True) if sym else "$") + whole.get_text(strip=True)
                if frac:
                    price += "." + frac.get_text(strip=True)
        data["价格"] = clean_text(price)

        # ---------- 评分 ----------
        rating_el = (
            soup.select_one("span[data-hook='rating-out-of-text']")
            or soup.select_one("i[data-hook='average-star-rating'] span")
            or soup.select_one("span.a-icon-alt")
        )
        data["评分"] = clean_text(rating_el.get_text(strip=True) if rating_el else None)

        # ---------- review 数量 ----------
        rc_el = (
            soup.select_one("#acrCustomerReviewText")
            or soup.select_one("span#acrCustomerReviewText")
            or soup.select_one("[data-hook='total-review-count']")
            or soup.select_one("#acrPopover .a-size-base")
        )
        data["review数量"] = clean_text(rc_el.get_text(strip=True) if rc_el else None)

        # ---------- 店铺名称 + 是否FBA ----------
        seller = "—"
        ships_from = "—"

        # 新版 tabular buybox
        for block in soup.select("#tabular-buybox .tabular-buybox-container, #tabular-buybox .tabular-buybox-text-row"):
            label_el = block.select_one(".tabular-buybox-label")
            text_el  = block.select_one(".tabular-buybox-text")
            if not label_el or not text_el:
                continue
            label = label_el.get_text(strip=True).lower()
            value = clean_text(text_el.get_text(strip=True))
            if "sold" in label and seller == "—":
                seller = value
            elif "ships" in label and ships_from == "—":
                ships_from = value

        # 旧式两行文本
        if seller == "—" or ships_from == "—":
            for box_sel in ["#shipsFromSoldBy_feature_div", "#desktop_buybox", "#rightCol", "#buybox_feature_div"]:
                box = soup.select_one(box_sel)
                if not box:
                    continue
                # Ships from
                if ships_from == "—":
                    lab = box.find(string=re.compile(r'^\s*Ships\s*from\s*$', re.I))
                    if lab:
                        row = lab.find_parent() or box
                        cand = row.find_next(lambda tag: tag.name in ["a", "span", "div"] and clean_text(tag.get_text()))
                        if cand:
                            val = clean_text(cand.get_text())
                            if val.lower() != "ships from":
                                ships_from = val
                # Sold by
                if seller == "—":
                    lab = box.find(string=re.compile(r'^\s*Sold\s*by\s*$', re.I))
                    if lab:
                        row = lab.find_parent() or box
                        cand = row.find_next(lambda tag: tag.name in ["a", "span", "div"] and clean_text(tag.get_text()))
                        if cand:
                            val = clean_text(cand.get_text())
                            if val.lower() != "sold by":
                                seller = val

                # 块内兜底
                if ships_from == "—":
                    m1 = re.search(r"Ships\s*from\s+([A-Za-z0-9 &\-]+)", box.get_text(" ", strip=True), re.I)
                    if m1:
                        ships_from = clean_text(m1.group(1))
                if seller == "—":
                    m2 = re.search(r"Sold\s*by\s+(.+?)(?:\s+and|\s+\.|$)", box.get_text(" ", strip=True), re.I)
                    if m2:
                        seller = clean_text(m2.group(1))

        # merchant-info 兜底
        if seller == "—":
            mi = soup.select_one("#merchant-info")
            if mi:
                m = re.search(r"Sold\s*by\s+(.+?)(?:\s+and|\s+\.|$)", mi.get_text(" ", strip=True), re.I)
                if m:
                    seller = clean_text(m.group(1))

        data["店铺名称"] = seller

        # 是否FBA
        is_fba = "否"
        if ships_from != "—" and "amazon" in ships_from.lower():
            is_fba = "是"
        else:
            blob = " ".join([
                soup.select_one("#merchant-info").get_text(" ", strip=True) if soup.select_one("#merchant-info") else "",
                soup.select_one("#tabular-buybox").get_text(" ", strip=True) if soup.select_one("#tabular-buybox") else "",
                soup.select_one("#shipsFromSoldBy_feature_div").get_text(" ", strip=True) if soup.select_one("#shipsFromSoldBy_feature_div") else "",
                soup.select_one("#desktop_buybox").get_text(" ", strip=True) if soup.select_one("#desktop_buybox") else "",
            ]).lower()
            if any(k in blob for k in ["fulfilled by amazon", "ships from amazon", "dispatched by amazon", "delivered by amazon"]):
                is_fba = "是"
        data["是否FBA"] = is_fba

        # ---------- 类目&排名 ----------
        bsr = "—"
        for sel in ["#detailBullets_feature_div", "#productDetails_detailBullets_sections1", "#prodDetails"]:
            node = soup.select_one(sel)
            if not node:
                continue
            text = node.get_text(" ", strip=True)
            mm = re.search(r"Best\s*Sellers?\s*Rank\s*:?\s*(.+?)(?:Date First Available|Customer Reviews|ASIN|$)", text, flags=re.I)
            if mm:
                bsr = clean_text(mm.group(1))
                break
        if bsr == "—":
            crumbs = [a.get_text(strip=True) for a in soup.select("#wayfinding-breadcrumbs_feature_div a")]
            if crumbs:
                bsr = " / ".join([c for c in crumbs if c])
        data["类目&排名"] = bsr

        # ---------- review 情况 ----------
        rv = (
            soup.select_one("div[data-hook='review'] span[data-hook='review-title'] span")
            or soup.select_one("div[data-hook='review'] span[data-hook='review-body'] span")
        )
        if rv:
            txt = rv.get_text(strip=True)
            data["review情况"] = clean_text(txt[:120] + ("..." if len(txt) > 120 else ""))
        else:
            data["review情况"] = "—"

        return data

    except Exception as e:
        print(f"[ERROR] {url} 抓取失败：{e}")
        return None

# ============ 最终简化+去重版店铺名称清洗模块 ============
def normalize_seller_name(name: str) -> str:
    """
    店铺名称清洗逻辑：
    - 去除前后空格
    - 如果包含 "Sold by"（不区分大小写），截断保留前部分
    - 去掉重复子串（如 "Conglin AU Conglin AU" → "Conglin AU"）
    """
    if not name or name == "—":
        return "—"

    s = name.strip()
    # 遇到 Sold by 就截断
    m = re.search(r"(?i)\bSold\s*by\b", s)
    if m:
        s = s[:m.start()]

    # 去除多余空格和标点
    s = s.strip(" .-–")

    # 判断重复（整串重复两遍的情况）
    parts = s.split()
    half = len(parts) // 2
    if len(parts) % 2 == 0 and parts[:half] == parts[half:]:
        s = " ".join(parts[:half])

    return s if s else "—"


def apply_seller_cleanup(rows):
    """就地清洗 rows 里的“店铺名称”字段"""
    for r in rows:
        if "店铺名称" in r:
            r["店铺名称"] = normalize_seller_name(r.get("店铺名称", "—"))


# ============ 生成带图片的 Excel ============
def save_xlsx_with_images(rows, xlsx_path="firemaple_playwright.xlsx"):
    """
    将抓取结果写入 .xlsx，并把“产品图片”嵌入首列缩略图。
    会尝试下载图片，失败则留空。
    """
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Fire-Maple AU"

    headers = ["产品图片","链接","亚马逊ASIN","价格","类目&排名","评分","店铺名称","是否FBA","review数量","review情况"]
    ws.append(headers)

    # 设置列宽，行高（首列放缩略图）
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 42
    for col_idx in range(3, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20

    # 下载图片用的简单 headers
    http_headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120 Safari/537.36"
    }

    row_idx = 2
    for row in rows:
        # 先写文本数据（图片列留空，稍后插入）
        ws.append([
            "",  # 图片稍后插入
            row.get("链接",""),
            row.get("亚马逊ASIN",""),
            row.get("价格",""),
            row.get("类目&排名",""),
            row.get("评分",""),
            row.get("店铺名称",""),
            row.get("是否FBA",""),
            row.get("review数量",""),
            row.get("review情况",""),
        ])

        # 拉取图片并缩略
        img_url = row.get("产品图片")
        if img_url and img_url != "—":
            try:
                r = requests.get(img_url, headers=http_headers, timeout=10)
                r.raise_for_status()
                img_bytes = io.BytesIO(r.content)
                with PILImage.open(img_bytes) as im:
                    im = im.convert("RGB")
                    im.thumbnail((120, 120))  # 控制缩略图大小
                    buf = io.BytesIO()
                    im.save(buf, format="JPEG", quality=85)
                    buf.seek(0)
                xl_img = XLImage(buf)
                xl_img.width, xl_img.height = im.size
                anchor = f"A{row_idx}"
                ws.add_image(xl_img, anchor)
                ws.row_dimensions[row_idx].height = 95  # 行高稍微大一点
            except Exception:
                # 下载失败就留空
                pass

        row_idx += 1

    from openpyxl.styles import Alignment
    for col in "BCDEFGHIJ":
        for r in range(1, row_idx):
            ws[f"{col}{r}"].alignment = Alignment(vertical="center", wrap_text=True)

    wb.save(xlsx_path)
    print(f"[DONE] 已生成带图片的 Excel：{xlsx_path}")

# ============ 主流程 ============
async def main():
    # 读取链接
    with open("urls.txt", "r", encoding="utf-8") as f:
        urls = [line.strip() for line in f if line.strip()]

    results = []
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        context = await browser.new_context(locale="en-AU", viewport={"width": 1280, "height": 900})
        page = await context.new_page()

        # 手动设置地址
        await set_au_delivery_address(page)

        for url in tqdm(urls, desc="抓取进度", unit="item"):
            data = await fetch_product(page, url)
            if data:
                results.append(data)
            await asyncio.sleep(2 + (random.random() * 2))

        await browser.close()

    # 店铺名称清洗（新增）
    apply_seller_cleanup(results)

    # 输出 CSV
    if results:
        df = pd.DataFrame(
            results,
            columns=[
                "产品图片",
                "链接",
                "亚马逊ASIN",
                "价格",
                "类目&排名",
                "评分",
                "店铺名称",
                "是否FBA",
                "review数量",
                "review情况",
            ],
        )
        csv_path = "firemaple_playwright.csv"
        xlsx_path = "firemaple_playwright.xlsx"
        df.to_csv(csv_path, index=False, encoding="utf-8-sig")
        print(f"[DONE] 共保存 {len(df)} 条到 CSV：{csv_path}")

        # 生成带图片的 Excel
        save_xlsx_with_images(results, xlsx_path=xlsx_path)
    else:
        print("[ERROR] 没有成功抓取到任何商品信息。")

if __name__ == "__main__":
    asyncio.run(main())
